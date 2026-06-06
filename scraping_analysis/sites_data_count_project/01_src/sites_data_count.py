#!/usr/bin/env python
# coding: utf-8

# In[86]:


import openpyxl as opxl
import requests
from bs4 import BeautifulSoup
from janome.tokenizer import Tokenizer    # janome.tokenizer.Tokenaizerをインポート
from collections import Counter
from time import sleep
from pathlib import Path
from janome.analyzer import Analyzer
from janome.charfilter import *
from janome.tokenfilter import *
import sqlite3
import os
import csv
import tkinter as tk 
from tkinter import messagebox 


# In[87]:


BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "03_data" / "dictionaries"
OUTPUT_DIR = BASE_DIR / "03_data" / "output"
REQUEST_TIMEOUT = 10

### INPUT
TMP_PATH = INPUT_DIR / "tmp_janome_dic.csv"             # 操作用のCSVを一時作成。元のCSVは操作しない
CSV_PATH = INPUT_DIR / "janome_dic.csv"                 # CSVの読み込み・書き出し先（同じファイル）
DB_PATH  = INPUT_DIR / "janome_dic.db"                  # SQLiteデータベース
ADD_DICT_PATH = INPUT_DIR / "add_dicword.txt"           # 辞書DBに追加するテキストファイル
NG_WORD_PATH = INPUT_DIR / "ng_word.txt"                # 追加するNGテキストファイル

### OUTPUT
SAVE_EXCEL_PATH = OUTPUT_DIR / "sites_data_count.xlsx"  # 結果書き込み用EXCELBOOK
SAVE_DB_PATH = OUTPUT_DIR / "sites_data_count.db"       # 結果書き込み用DB

STOP_WORDS = []                                         # NGワード


def env_flag(name):
    return os.getenv(name, "").lower() == "true"


# In[88]:


# 環境変数で切り替え可能なオプション機能
# 形態素解析のNGワード追加
def set_ng_word(ng_word_text):
    STOP_WORDS = []
    with open(ng_word_text, "r", encoding="utf-8")  as  file1:
        for data1 in file1:
            add_word = data1.strip().replace(" ","").split(",")
            add_word = [x for x in add_word if x != ""]
            STOP_WORDS.append(add_word)
    STOP_WORDS = [x for row in STOP_WORDS for x in row]
    return STOP_WORDS


# In[89]:


def get_conn():
# SQLアクセス処理
    try :
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn
    except sqlite3.Error as e:
        raise RuntimeError(f"DBに接続できませんでした: {e}")


# In[90]:


# 環境変数で切り替え可能なオプション機能
# 辞書CSV⇒辞書DBに変換
def convert_CSV2SQL(rfile_path):
    num_column = 13        # 辞書の項目数
    conn = get_conn()
    cur  = conn.cursor()
    
    cur.execute('DROP TABLE IF EXISTS dictionary')
    cur.execute('''
        CREATE TABLE dictionary (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            surface       TEXT NOT NULL,
            left_id       INTEGER,
            right_id      INTEGER,
            cost          INTEGER,
            pos1          TEXT,
            pos2          TEXT,
            pos3          TEXT,
            pos4          TEXT,
            conj_type     TEXT,
            conj_form     TEXT,
            base_form     TEXT,
            reading       TEXT,
            pronunciation TEXT
        )
    ''')
    
    count = 0
    with open(rfile_path, encoding='utf-8') as f:
        for row in csv.reader(f):
            if len(row) < num_column:
                continue
            cur.execute('''
                INSERT INTO dictionary
                (surface, left_id, right_id, cost,
                 pos1, pos2, pos3, pos4,
                 conj_type, conj_form, base_form, reading, pronunciation)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
            ''', row[:num_column])
            count += 1
    
    conn.commit()
    conn.close()


# In[91]:


# 環境変数で切り替え可能なオプション機能
# 辞書DB⇒辞書CSVに変換
def convert_SQL2CSV(wfile_path):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute('''
        SELECT surface, left_id, right_id, cost,
               pos1, pos2, pos3, pos4,
               conj_type, conj_form, base_form, reading, pronunciation
        FROM dictionary
        ORDER BY surface
    ''')
    rows = cur.fetchall()
    conn.close()
    
    with open(wfile_path, 'w', encoding='utf-8', newline='') as f:
        csv.writer(f).writerows([list(r) for r in rows])
    return True


# In[92]:


def copyCSV(rfile_path, wfile_path):
    # 辞書データを一時作業用CSVに保存
    rows = []
    with open(rfile_path, 'r', encoding='utf-8') as f:
        rows = list(csv.reader(f))

    with open(wfile_path, 'w', encoding='utf-8', newline='') as f:
        csv.writer(f).writerows(rows)


# In[93]:


def delDict():
    # 一時作業用CSVを削除
    if os.path.exists(TMP_PATH):
        os.remove(TMP_PATH)


# In[94]:


# 環境変数で切り替え可能なオプション機能
# テキストファイルから辞書に追加
def addDict(url_add_txt):
    words = []
    with open(url_add_txt, "r", encoding="utf-8")  as  file1:
        for skip in range(3):
            next(file1, None)        # 最初の3行はコメントのため、読み飛ばす
        for data1 in file1:
            add_word = data1.strip().replace("\"",'')
            add_word = tuple(add_word.split(","))
            if len(add_word) == 3:
                words.append(add_word)
    conn = get_conn()
    cur  = conn.cursor()
    
    for surface, reading, pronunciation in words:
        cur.execute('SELECT id FROM dictionary WHERE surface=? AND reading=?', (surface, reading))
        if cur.fetchone():
            print(f'スキップ（既存）: {surface}')
            continue
        cur.execute('''
            INSERT INTO dictionary
            (surface, left_id, right_id, cost,
             pos1, pos2, pos3, pos4,
             conj_type, conj_form, base_form, reading, pronunciation)
            VALUES (?,1288,1288,4000,'名詞','固有名詞','一般','*','*','*',?,?,?)
        ''', (surface, surface, reading, pronunciation))
        print(f'追加: {surface}')

    conn.commit()
    conn.close()


# In[95]:


def checkFileWritable(filename=SAVE_EXCEL_PATH):
    """Excelファイルが書き込み可能な状態かを事前にチェックする"""
    
    # ファイルが存在する場合のみチェック（新規作成の場合はスキップ）
    if os.path.exists(filename):
        try:
            # 追記モードで開いて即閉じる（実際に書き込めるか確認）
            with open(filename, "a"):
                pass
        except PermissionError:
            raise PermissionError(f"エラー: '{filename}' が他のアプリで開かれています。Excelを閉じてから再実行してください。")
            return False
        except OSError as e:
            raise OSError(f"エラー: '{filename}' へのアクセスに失敗しました。({e})")
            return False
    
    return True


# In[96]:


# collectURL()関数
def collectURL(url):
    
    # requestで記事一覧ページのHTMLを取得
    list_response = requests.get(url, timeout=REQUEST_TIMEOUT)
    list_response.raise_for_status()
    
    # soupからaタグ（href属性あり）を全部リストで取得
    soup = BeautifulSoup(list_response.text, 'html.parser')
    all_links = soup.find_all("a", href=True)
    
    # aタグから記事のURLを抽出・絞り込み
    article_urls = []
    count = 0

    for links in all_links:
        href = links["href"]
        if href.startswith("https://gomafree-fun.site/?p="):
            if href not in article_urls:  # 重複防止
                article_urls.append(href)
                count = count + 1
                
    if count == 0:
        print("エラー:記事一覧に記事がありません")
        raise ValueError("エラー:記事一覧に記事がありません")
        return []
                
    print(f"記事合計:{count}本")

    return article_urls

    # 記事のURLは("https://gomafree-fun.site/?p=")ですると余計なものが入らない


# In[97]:


# getText()関数
def getText(article_urls):
    
    # 空の配列を用意する
    article_texts = []
    
    # for文を使って各記事のURLから本文のみを抽出
    # requestsを使って記事のHTMLを丸ごと取得
    # soupを使って記事本文のみを抽出(本文のdivはclass_="entry-content")
    # 空の配列に本文を追加する
    for article_url in article_urls:
        response = requests.get(article_url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        article_soup = BeautifulSoup(response.text, "html.parser")
        article_text = article_soup.find('div', class_="entry-content")


         # 例外処理：特定の記事だけクラスが見つからない場合
        if article_text is None:
            print(f"スキップ:{article_url}の本文が取得できませんでした")
            sleep(1)
            continue
        
        article_texts.append(article_text.get_text(strip=True))
        sleep(1)

    return article_texts


# In[98]:


def anlzNoun(goma_list, stop_words):
    """本文リストを形態素解析し、品詞別の単語リストを返す。"""
    # iPhone 14 Pro など、英数字の間に入った空白を解析前に詰める
    char_filters = [RegexReplaceCharFilter(r"([a-z])\s([0-9])", r"\1\2")
                  , RegexReplaceCharFilter(r"([0-9])\s([a-z])", r"\1\2")]
    t = Tokenizer(str(TMP_PATH), udic_enc="utf8")
    
# 文章を最小単位の単語に分割する
# 単語と、品詞を抽出(保持はしない)

    # 空のリストを用意
    nouns = []
    verbs = []
    adjectives = []
    adverbs = []

# Analyzerを使用する
    a = Analyzer(char_filters = char_filters, tokenizer=t, token_filters=[CompoundNounFilter()])
    
# 単語の品詞、細分類の情報をリスト化
# 先頭の要素(品詞)のみを抜き出す。
    for text in goma_list:
        text = text.lower()                  #英語の大文字小文字を統一
        text = text.replace(".", "")
        tokens = a.analyze(text)
        #1文字は除外
        for token in tokens:
            if len(token.surface) <= 1:      # 1文字を除外
                continue
            if token.surface in stop_words:  # ストップワードを除外
                continue
            part_of_speech = token.part_of_speech.split(",")
            major = part_of_speech[0]
            minor = part_of_speech[1]
        # 抜き出した要素が名詞か
        # (名詞)下記、代名詞/数字除外の判定処理を行う
        #代名詞は文字型のメソッドstr.find("代名詞")かつ、str.find("数字")が-1を返すか？
            if major == "名詞":
                if minor not in ["代名詞", "数"]:
                    nouns.append(token.surface)
                else:
                    continue
            elif major == "動詞":
                if minor == "自立":
                    verbs.append(token.surface)
            elif major == "形容詞":
                if minor == "自立":
                    adjectives.append(token.surface)
            elif major == "副詞":
                adverbs.append(token.surface)
            else:
                continue
#  ⇒(-1を返す)上記で用意したリストに単語を追加する
#  ⇒(それ以外)無視します:continue
    return nouns, verbs, adjectives, adverbs


# In[99]:


def revSortCount(nouns):
# カウンタライブラリを使用し、リスト中に同じ単語が何回登場したかカウント(←重複しない)
    noun_counts = Counter(nouns)
# カウンタクラスから辞書型に変換
    word_dict = dict(noun_counts)
# 降順にソートした後、sorted関数で、keyと、valueのタプルを返す
    word_dict = sorted(
    word_dict.items()       # 対象の辞書
    , key=lambda x : x[1]   # 並び替えの基準(key)を辞書の値
    , reverse=True          # 降順で並び替え
)
    return word_dict


# In[100]:


def saveExcel(noun_count, verb_count, adj_count, adv_count):
    # Workbook作成(opxl.Workbook()使用)
    book = opxl.Workbook()
    sheets = [
        ("名詞", noun_count),
        ("動詞", verb_count),
        ("形容詞", adj_count),
        ("副詞", adv_count),
    ]

    for sheet_name, word_dict in sheets:
        sheet = book.create_sheet(sheet_name)
        sheet["A1"] = "単語"
        sheet["B1"] = "出現回数"
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 12
        sheet.freeze_panes = "A2"
        
        # 30以上か？
        if len(word_dict) > 30:
            num_word = 30             # 30にクリップ
        else:
            num_word = len(word_dict)
    
    # 引数の単語(key)と、カウント値(value)を先頭からサイズ分(最大30)書き込み
        for i in range(1, num_word + 1):
            sheet["A" + str(i + 1)] = word_dict[i-1][0]
            sheet["B" + str(i + 1)] = word_dict[i-1][1]
    # sites_data_count.xlsxで保存(book1.save()使用)

    # openpyxlはデフォルトで"Sheet"が作られるので削除
    del book["Sheet"]
    
    try:
        book.save(SAVE_EXCEL_PATH)
        print("保存成功: sites_data_count.xlsx")
        print("実行完了")
    except PermissionError:
        raise PermissionError("エラー: ファイルが他のアプリで開かれています。Excelを閉じてから再実行してください。")
    except OSError as e:
        raise OSError(f"エラー: ファイルの保存に失敗しました。({e})")
    except Exception as e:
        raise Exception(f"予期しないエラーが発生しました。({e})")


# In[101]:


def saveDB(word_dict, dbname, table_name="word_count"):

    # 上位30件に絞る
    top30 = word_dict[:30]

    # カレントディレクトリにdbがあれば接続。なければ作成。
    conn = sqlite3.connect(dbname)

    # sqliteを操作するカーソルを作成 データベースのどこ(行)でを管理
    cur = conn.cursor()

    # テーブル構造
    cur.execute(f"""
    CREATE TABLE IF NOT EXISTS {table_name} (
        word TEXT PRIMARY KEY,
        count INTEGER NOT NULL
    )
    """)

    # 既存データをクリア
    cur.execute(f"DELETE FROM {table_name}")

    # データ挿入
    for word, count in top30:
        cur.execute(
            f"INSERT OR REPLACE INTO {table_name} (word, count) VALUES (?, ?)",
            (word, count)
        )

    conn.commit()
    conn.close()


# In[102]:


def preProcessEnb():
############## cmdから環境変数でセットするオプション機能 #############
    # f_ng_word: 形態素解析で除外したいワードをテキストファイルから取り込むか False:無効(デフォルト), True:有効
    if env_flag("f_ng_word"):
        STOP_WORDS.extend(set_ng_word(NG_WORD_PATH))

    # f_use_db : DBを辞書として使用するか切り替える False:未使用(デフォルト), True:使用
    if env_flag("f_use_db"):
        # f_csv2db : 辞書csvの内容で、辞書DBを上書き False:未使用(デフォルト), True:使用
        if env_flag("f_csv2db"):
            convert_CSV2SQL(CSV_PATH)
        elif not os.path.exists(DB_PATH):
            print("辞書DBが存在しません。f_csv2db=True でCSVから作成できます。")
            return False

        # f_add_dic: DBを辞書とする際、テキストファイルから単語を登録するか False:無効(デフォルト), True:有効
        if env_flag("f_add_dic") or env_flag("f_add_dict"):
            addDict(ADD_DICT_PATH)
        convert_SQL2CSV(TMP_PATH)

    # デフォルトはCSVを辞書として使用する
    else :
        if not os.path.exists(CSV_PATH):
            print("辞書CSVが存在しません。")
            return False
        copyCSV(CSV_PATH, TMP_PATH)
    return True
#################################################################


# In[104]:


def run_analysis():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if not checkFileWritable(SAVE_EXCEL_PATH):
        return False
    if not preProcessEnb():
        return False
    url = "https://gomafree-fun.site/?page_id=278"
    collecturl = collectURL(url)
    gettext = getText(collecturl)
    nouns, verbs, adjectives, adverbs = anlzNoun(gettext, STOP_WORDS)  # 追加

    # 品詞毎、単語の出現回数をカウントし、降順にソート
    noun_count = revSortCount(nouns)
    verb_count = revSortCount(verbs)
    adj_count  = revSortCount(adjectives)
    adv_count  = revSortCount(adverbs)

    # 品詞毎、単語と、出現回数をDBに保存
    saveDB(noun_count, SAVE_DB_PATH, table_name="noun_count")
    saveDB(verb_count, SAVE_DB_PATH, table_name="verb_count")
    saveDB(adj_count, SAVE_DB_PATH, table_name="adj_count")
    saveDB(adv_count, SAVE_DB_PATH, table_name="adv_count")
        
    saveExcel(noun_count, verb_count, adj_count, adv_count)                  # 各品詞の単語と、出現回数をExcelに保存
    delDict()                                                                # 一時保存用の辞書ファイル削除
    return True


# In[105]:


def show_main_window():
############## cmdから環境変数でセットするオプション機能 #############
    if env_flag("f_display_ui"):
        # ウィンドウ作成
        root = tk.Tk()
        root.title("頻出ワード解析")
        root.geometry("280x180")
        root.configure(bg="#f5f5f5")
        root.resizable(False, False)
    
        label = tk.Label(
            root,
            text="頻出ワード解析",
            bg="#f5f5f5",
            fg="#333333",
            font=("Yu Gothic", 13, "bold")
        )
        label.pack(pady=(30, 10))
    
        def start():
            btn.destroy()
            label.config(text="実行中...")
            root.update()
            try:
                if not run_analysis():
                    return
                label.config(text="完了！")
            except PermissionError as e:
                messagebox.showerror("エラー", str(e))
            except OSError as e:
                messagebox.showerror("エラー", str(e))
            except ValueError as e:
                messagebox.showerror("エラー", str(e))
            except requests.exceptions.RequestException as e:
                messagebox.showerror("エラー", f"通信エラーが発生しました。({e})")
            except Exception as e:
                messagebox.showerror("エラー", str(e))
            finally :
                close_btn = tk.Button(
                    root,
                    text="閉じる",
                    command=root.destroy,
                    bg="#333333",
                    fg="#ffffff",
                    font=("Yu Gothic", 10),
                    relief=tk.FLAT,
                    padx=20,
                    pady=6,
                    cursor="hand2"
                )
                close_btn.pack(pady=10)
    
        btn = tk.Button(
            root,
            text="開始",
            bg="#333333",
            fg="#ffffff",
            font=("Yu Gothic", 10),
            relief=tk.FLAT,
            padx=20,
            pady=6,
            cursor="hand2",
            command=start
        )
        btn.pack()
        root.mainloop()
#################################################################

# デフォルトはGUIなし
    else:
        if not run_analysis():
            return


# In[82]:


def main():
    show_main_window();


# In[83]:


if __name__ == "__main__":
    main()
